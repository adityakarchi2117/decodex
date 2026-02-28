"""
Build Final STABILIS Submission Package
=======================================
Creates:
  1. Final_STABILIS_Submission.xlsx  (6 sheets, clean, formatted)
  2. final_submission/               (clean folder structure)
  3. Console output: file keep/discard list, chart list, completeness check
"""

import json, shutil, os
from pathlib import Path
import pandas as pd
import numpy as np

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────────────────────
#  PATHS
# ─────────────────────────────────────────────────────────────────────────────
BASE        = Path(".")
STAGE1_OUT  = BASE / "outputs"
STAGE1_TBL  = STAGE1_OUT / "tables"
STAGE1_CHR  = STAGE1_OUT / "charts"
STAGE2_VAL  = BASE / "stage2_jury_proof_validation_pack"
STAGE2_SHK  = BASE / "stage2_shock_pack"
STAGE3_OUT  = BASE / "stage3_optimization_pack"
STAGE3_TBL  = STAGE3_OUT / "tables"
STAGE3_CHR  = STAGE3_OUT / "charts"

SUBMIT_DIR  = BASE / "final_submission"
SUBMIT_CHR  = SUBMIT_DIR / "charts"
SUBMIT_XL   = SUBMIT_DIR / "Final_STABILIS_Submission.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
#  STEP 1 — LOAD ALL REQUIRED DATA
# ─────────────────────────────────────────────────────────────────────────────

def load_stage1_baseline():
    """Stage 1 baseline summary table."""
    with open(STAGE1_OUT / "model_metrics.json", encoding="utf-8") as f:
        s1 = json.load(f)

    baseline = s1["baseline_kpis"]
    fragility = s1["fragility_metrics"]
    modeling = s1["modeling_metrics"]
    ci = pd.read_csv(STAGE1_TBL / "bootstrap_ci_y1.csv")

    rows = [
        ["ERPU Mean",                  f"{baseline['erpu_mean']:,.2f}"],
        ["ERPU Median",                f"{baseline['erpu_median']:,.2f}"],
        ["Total Users",                f"{baseline['total_users']:,}"],
        ["Total Baskets",              f"{baseline['total_baskets']:,}"],
        ["Net Revenue (Y1)",           f"{baseline['net_revenue']:,.2f}"],
        ["Return Rate (%)",            f"{baseline['return_rate_pct']:.2f}%"],
        ["Top 10% User Share (Net)",   f"{fragility['top10_user_share_net']:.2%}"],
        ["Top 1% User Share (Net)",    f"{fragility['top1_user_share_net']:.2%}"],
        ["Gini Coefficient",           f"{fragility['gini_user_net_y1']:.4f}"],
        ["Module A AUC",               f"{modeling['repeat_purchase_auc']:.4f}"],
        ["Module A Brier Score",       f"{modeling['brier_score']:.4f}"],
        ["Module A Top-Decile Lift",   f"{modeling['top_decile_lift']:.4f}"],
        ["Module C R-squared",         f"{modeling['basket_value_model_r2']:.4f}"],
        ["NegBin Dispersion",          f"{modeling['poisson_dispersion']:.4f}"],
    ]

    # Add bootstrap CIs
    for _, row in ci.iterrows():
        rows.append([
            f"Bootstrap CI: {row['metric']}",
            f"{row['point_estimate']:.4f}  [{row['ci_lo']:.4f}, {row['ci_hi']:.4f}]"
        ])

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def load_stage2_validation():
    """Stage 2 validation summary."""
    dash = pd.read_csv(STAGE2_VAL / "train_vs_validation_dashboard.csv")
    conc = pd.read_csv(STAGE2_VAL / "concentration_stability.csv")

    # Clean dashboard: select key metrics only
    key_metrics = [
        "AUC", "Brier Score", "Log Loss",
        "ECE (Calibration Error)", "Top Decile Lift",
        "Mean ERPU", "Top 10% User Share", "Gini Coefficient"
    ]
    dash_clean = dash[dash["metric"].isin(key_metrics)].copy()
    dash_clean = dash_clean.reset_index(drop=True)

    # Format numbers
    for col in ["train", "validation", "delta"]:
        dash_clean[col] = dash_clean[col].apply(
            lambda x: f"{x:.4f}" if abs(x) < 10 else f"{x:,.2f}"
        )
    dash_clean["delta_pct"] = dash_clean["delta_pct"].apply(lambda x: f"{x:.2f}%")

    dash_clean.columns = ["Metric", "Train (Y1)", "Validation (Y2)", "Delta", "Delta %"]

    # Concentration stability
    conc_rows = []
    for _, r in conc.iterrows():
        conc_rows.append([
            r["dataset"],
            f"{r['top10_pct_share']:.2%}",
            f"{r['gini_coefficient']:.4f}",
            f"{r['erpu_mean']:,.2f}",
            f"{r['erpu_median']:,.2f}",
            f"{r['mean_median_ratio']:.2f}"
        ])
    conc_df = pd.DataFrame(conc_rows, columns=[
        "Dataset", "Top 10% Share", "Gini", "ERPU Mean", "ERPU Median", "Mean/Median Ratio"
    ])

    return dash_clean, conc_df


def load_stage3_scoring():
    """Stage 3 test scoring table — cleaned columns only."""
    df = pd.read_csv(STAGE3_TBL / "test_user_scores.csv")

    keep = [
        "customerid",
        "pred_purchase_prob",
        "pred_frequency",
        "pred_basket_value",
        "pred_return_rate",
        "E_gross",
        "E_return",
        "pred_erpu",
        "actual_erpu",
        "return_risk_decile",
        "targeted",
    ]
    df_clean = df[keep].copy()
    df_clean.columns = [
        "UserID",
        "Purchase Probability",
        "Expected Frequency",
        "Expected Basket Value",
        "Expected Return Rate",
        "Expected Gross Revenue",
        "Expected Return",
        "Expected Net Revenue (ERPU)",
        "Actual ERPU",
        "Return Risk Decile",
        "Target Flag",
    ]
    return df_clean


def load_optimization_summary():
    """Constraint compliance + optimization KPIs."""
    compliance = pd.read_csv(STAGE3_TBL / "constraint_compliance.csv")

    with open(STAGE3_OUT / "model_metrics.json", encoding="utf-8") as f:
        s3 = json.load(f)

    kpi_rows = [
        ["Total Expected Net Revenue",   f"{s3['total_expected_net_revenue']:,.2f}"],
        ["Targeted User Count",          f"{s3['n_targeted_users']:,}"],
        ["Total Scored Users",           f"{s3['n_total_users']:,}"],
        ["% Targeted",                   f"{s3['pct_targeted']:.2f}%"],
        ["Mean ERPU (Targeted)",         f"{s3['mean_erpu_targeted']:,.2f}"],
        ["Mean ERPU (Non-Targeted)",     f"{s3['mean_erpu_non_targeted']:,.2f}"],
        ["Optimization Method",           s3["optimization_method"]],
        ["Lambda 1 (Return Risk)",       str(s3["lambda1"])],
        ["Lambda 2 (Concentration)",     str(s3["lambda2"])],
        ["Return Risk Exposure",         f"{s3['constraints']['return_risk']['actual']:.1%}"],
        ["Top 10% Revenue Share",        f"{s3['constraints']['concentration']['stage3_top10_share']:.2%}"],
        ["Concentration Delta vs S1",    f"{s3['constraints']['concentration']['relative_increase']:.2%}"],
        ["Stage 1 Top 10% Baseline",     f"{s3['stage1_baseline']['top10_user_share']:.2%}"],
    ]

    # ERPU decomposition
    decomp = s3["erpu_decomposition_targeted"]
    kpi_rows += [
        ["", ""],
        ["--- ERPU Decomposition (Targeted) ---", ""],
        ["Mean P(purchase)",             f"{decomp['mean_purchase_prob']:.4f}"],
        ["Mean E[frequency]",            f"{decomp['mean_frequency']:.4f}"],
        ["E[basket value]",              f"{decomp['mean_basket_value']:,.2f}"],
        ["Return Rate",                  f"{decomp['return_rate']:.4f}"],
        ["Mean E[gross]",                f"{decomp['mean_E_gross']:,.2f}"],
        ["Mean E[return]",               f"{decomp['mean_E_return']:,.2f}"],
        ["Mean ERPU",                    f"{decomp['mean_erpu']:,.2f}"],
    ]

    kpi_df = pd.DataFrame(kpi_rows, columns=["Metric", "Value"])
    return kpi_df, compliance


def load_tradeoff():
    """Trade-off table cleaned."""
    df = pd.read_csv(STAGE3_TBL / "tradeoff_table.csv")
    df_clean = df[["scenario", "n_targeted", "total_erpu", "mean_erpu",
                    "return_prone_pct", "top10_share_pct"]].copy()
    df_clean.columns = [
        "Target %", "N Targeted", "Total Revenue",
        "Mean ERPU", "Return Risk %", "Top 10% Share %"
    ]
    df_clean["Total Revenue"] = df_clean["Total Revenue"].apply(lambda x: f"{x:,.0f}")
    df_clean["Mean ERPU"] = df_clean["Mean ERPU"].apply(lambda x: f"{x:,.0f}")
    df_clean["Return Risk %"] = df_clean["Return Risk %"].apply(lambda x: f"{x:.2f}%")
    df_clean["Top 10% Share %"] = df_clean["Top 10% Share %"].apply(lambda x: f"{x:.2f}%")
    return df_clean


def load_sensitivity():
    """Sensitivity / fragility table cleaned."""
    df = pd.read_csv(STAGE3_TBL / "sensitivity_table.csv")
    df_clean = df[["scenario", "total_erpu", "return_pct",
                    "top10_share_pct", "revenue_impact_pct", "all_feasible"]].copy()
    df_clean.columns = [
        "Scenario", "Total Revenue", "Return Risk %",
        "Top 10% Share %", "Revenue Impact %", "All Feasible"
    ]
    df_clean["Total Revenue"] = df_clean["Total Revenue"].apply(lambda x: f"{x:,.0f}")
    df_clean["Return Risk %"] = df_clean["Return Risk %"].apply(lambda x: f"{x:.2f}%")
    df_clean["Top 10% Share %"] = df_clean["Top 10% Share %"].apply(lambda x: f"{x:.2f}%")
    df_clean["Revenue Impact %"] = df_clean["Revenue Impact %"].apply(lambda x: f"{x:.2f}%")
    df_clean["All Feasible"] = df_clean["All Feasible"].map({True: "YES", False: "NO"})
    return df_clean


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 2 — BUILD EXCEL WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

def style_sheet(ws, df, header_fill, header_font, border):
    """Apply formatting to a worksheet."""
    # Column widths
    for i, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df.iloc[:, i-1].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)

    # Header styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Data styling
    data_font = Font(size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.sheet_properties.tabColor = "1F4E79"


def build_workbook():
    """Build the final Excel with all 6 sheets."""
    SUBMIT_DIR.mkdir(parents=True, exist_ok=True)

    # Load all data
    s1_baseline = load_stage1_baseline()
    s2_dash, s2_conc = load_stage2_validation()
    s3_scoring = load_stage3_scoring()
    opt_kpis, compliance = load_optimization_summary()
    tradeoff = load_tradeoff()
    sensitivity = load_sensitivity()

    with pd.ExcelWriter(str(SUBMIT_XL), engine="openpyxl") as writer:
        # Sheet 1: Stage1_Baseline
        s1_baseline.to_excel(writer, sheet_name="Stage1_Baseline", index=False)

        # Sheet 2: Stage2_Validation (two tables)
        s2_dash.to_excel(writer, sheet_name="Stage2_Validation", index=False, startrow=0)
        # Add concentration stability below with a gap
        gap_row = len(s2_dash) + 3
        pd.DataFrame([["Concentration Stability Comparison"]], columns=[""]).to_excel(
            writer, sheet_name="Stage2_Validation", index=False,
            startrow=gap_row - 1, header=False
        )
        s2_conc.to_excel(writer, sheet_name="Stage2_Validation", index=False, startrow=gap_row)

        # Sheet 3: Stage3_Test_Scoring (full user-level table)
        s3_scoring.to_excel(writer, sheet_name="Stage3_Test_Scoring", index=False)

        # Sheet 4: Optimization_Summary (KPIs + constraint compliance)
        opt_kpis.to_excel(writer, sheet_name="Optimization_Summary", index=False, startrow=0)
        gap2 = len(opt_kpis) + 3
        pd.DataFrame([["Constraint Compliance"]], columns=[""]).to_excel(
            writer, sheet_name="Optimization_Summary", index=False,
            startrow=gap2 - 1, header=False
        )
        compliance.to_excel(writer, sheet_name="Optimization_Summary", index=False, startrow=gap2)

        # Sheet 5: Tradeoff_Table
        tradeoff.to_excel(writer, sheet_name="Tradeoff_Table", index=False)

        # Sheet 6: Sensitivity_Analysis
        sensitivity.to_excel(writer, sheet_name="Sensitivity_Analysis", index=False)

    # Apply formatting if openpyxl styles available
    if HAS_OPENPYXL:
        from openpyxl import load_workbook
        wb = load_workbook(str(SUBMIT_XL))

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        sheets_data = {
            "Stage1_Baseline": s1_baseline,
            "Stage2_Validation": s2_dash,
            "Stage3_Test_Scoring": s3_scoring,
            "Optimization_Summary": opt_kpis,
            "Tradeoff_Table": tradeoff,
            "Sensitivity_Analysis": sensitivity,
        }

        for name, df in sheets_data.items():
            style_sheet(wb[name], df, header_fill, header_font, thin_border)

        # Freeze panes on scoring sheet
        wb["Stage3_Test_Scoring"].freeze_panes = "A2"

        wb.save(str(SUBMIT_XL))

    print(f"  [OK] Workbook saved: {SUBMIT_XL}")
    return {
        "s1": s1_baseline, "s2_dash": s2_dash, "s2_conc": s2_conc,
        "s3": s3_scoring, "opt": opt_kpis, "compliance": compliance,
        "tradeoff": tradeoff, "sensitivity": sensitivity
    }


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 3 — CHART SELECTION & FOLDER STRUCTURE
# ─────────────────────────────────────────────────────────────────────────────

# Final chart manifest
FINAL_CHARTS = {
    # Stage 1
    "S1_lorenz_curve.png":              STAGE1_CHR / "01_lorenz_curve_y1.png",
    "S1_basket_log_distribution.png":   STAGE1_CHR / "02_basket_net_log_hist_y1.png",
    "S1_revenue_concentration.png":     STAGE1_CHR / "09_cumshare_users_net_y1.png",
    "S1_calibration_plot.png":          STAGE1_CHR / "13_moduleA_calibration_plot.png",
    # Stage 2
    "S2_calibration_curve.png":         STAGE2_VAL / "charts" / "calibration_plot.png",
    "S2_reliability_curve.png":         STAGE2_VAL / "charts" / "reliability_curve.png",
    "S2_gain_curve.png":                STAGE2_VAL / "charts" / "gain_curve.png",
    # Stage 3
    "S3_revenue_vs_return_risk.png":    STAGE3_CHR / "01_revenue_vs_return_risk.png",
    "S3_revenue_vs_concentration.png":  STAGE3_CHR / "02_revenue_vs_concentration.png",
    "S3_marginal_gain_budgets.png":     STAGE3_CHR / "03_marginal_gain_budgets.png",
    "S3_erpu_dist_targeted.png":        STAGE3_CHR / "04_erpu_distribution_targeted_vs_nontargeted.png",
    "S3_constraint_compliance.png":     STAGE3_CHR / "06_constraint_compliance_table.png",
    "S3_pred_vs_actual_erpu.png":       STAGE3_CHR / "07_pred_vs_actual_erpu.png",
}


def copy_charts():
    """Copy selected final charts to submission folder."""
    SUBMIT_CHR.mkdir(parents=True, exist_ok=True)
    copied = []
    missing = []
    for dest_name, src_path in FINAL_CHARTS.items():
        if src_path.exists():
            shutil.copy2(str(src_path), str(SUBMIT_CHR / dest_name))
            copied.append(dest_name)
        else:
            missing.append(f"{dest_name} <-- {src_path}")
    return copied, missing


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 4 — FILE CLASSIFICATION (KEEP / DISCARD)
# ─────────────────────────────────────────────────────────────────────────────

def classify_files():
    """Classify all generated artifacts as KEEP or DISCARD."""
    keep = []
    discard = []

    # ── Stage 1 ──────────────────────────────────────────
    s1_keep_tables = {
        "bootstrap_ci_y1.csv",
        "erpu_decomposition_y1.csv",
        "moduleA_calibration_deciles.csv",
    }
    s1_keep_charts = {
        "01_lorenz_curve_y1.png",
        "02_basket_net_log_hist_y1.png",
        "09_cumshare_users_net_y1.png",
        "13_moduleA_calibration_plot.png",
    }
    s1_keep_top = {
        "model_metrics.json",
        "Stage1_Summary.txt",
        "Stage1_Baseline_for_Stage2.json",
    }

    for f in (STAGE1_TBL).glob("*"):
        if f.name in s1_keep_tables:
            keep.append(("Stage1/tables", f.name, "Baseline CI / ERPU decomposition / calibration"))
        else:
            discard.append(("Stage1/tables", f.name, "Intermediate diagnostic"))

    for f in (STAGE1_CHR).glob("*"):
        if f.name in s1_keep_charts:
            keep.append(("Stage1/charts", f.name, "Final presentation chart"))
        else:
            discard.append(("Stage1/charts", f.name, "Intermediate histogram / diagnostic"))

    for f in STAGE1_OUT.glob("*"):
        if f.is_file():
            if f.name in s1_keep_top:
                keep.append(("Stage1", f.name, "Baseline reference"))
            elif f.name == "Stage1_Strategy_Frameworks.txt":
                discard.append(("Stage1", f.name, "Internal strategy notes"))

    # ── Stage 2 ──────────────────────────────────────────
    s2_val_keep = {
        "train_vs_validation_dashboard.csv",
        "concentration_stability.csv",
        "model_metrics.json",
    }
    s2_val_charts_keep = {
        "calibration_plot.png",
        "reliability_curve.png",
        "gain_curve.png",
    }

    for f in STAGE2_VAL.glob("*"):
        if f.is_file():
            if f.name in s2_val_keep:
                keep.append(("Stage2_Validation", f.name, "Validation summary"))
            else:
                discard.append(("Stage2_Validation", f.name, "Internal diagnostic / jury-proof artifact"))

    for f in (STAGE2_VAL / "charts").glob("*"):
        if f.name in s2_val_charts_keep:
            keep.append(("Stage2_Validation/charts", f.name, "Final presentation chart"))
        else:
            discard.append(("Stage2_Validation/charts", f.name, "Intermediate diagnostic plot"))

    # Stage 2 shock pack — all diagnostic, discard
    for f in STAGE2_SHK.glob("*"):
        if f.is_file():
            discard.append(("Stage2_Shock", f.name, "Shock-test diagnostic (internal)"))
    for f in (STAGE2_SHK / "charts").glob("*"):
        discard.append(("Stage2_Shock/charts", f.name, "Shock comparison chart (internal)"))

    # ── Stage 3 ──────────────────────────────────────────
    s3_keep_tables = {
        "test_user_scores.csv",
        "targeted_users.csv",
        "constraint_compliance.csv",
        "tradeoff_table.csv",
        "sensitivity_table.csv",
    }
    s3_discard_tables = {
        "revenue_vs_return_risk.csv",
        "revenue_vs_concentration.csv",
    }
    s3_keep_charts = {
        "01_revenue_vs_return_risk.png",
        "02_revenue_vs_concentration.png",
        "03_marginal_gain_budgets.png",
        "04_erpu_distribution_targeted_vs_nontargeted.png",
        "06_constraint_compliance_table.png",
        "07_pred_vs_actual_erpu.png",
    }

    for f in STAGE3_TBL.glob("*"):
        if f.name in s3_keep_tables:
            keep.append(("Stage3/tables", f.name, "Final submission artifact"))
        elif f.name in s3_discard_tables:
            discard.append(("Stage3/tables", f.name, "Chart source data (redundant)"))
        else:
            discard.append(("Stage3/tables", f.name, "Unknown/intermediate"))

    for f in STAGE3_CHR.glob("*"):
        if f.name in s3_keep_charts:
            keep.append(("Stage3/charts", f.name, "Final presentation chart"))
        else:
            discard.append(("Stage3/charts", f.name, "Secondary visualization"))

    for f in STAGE3_OUT.glob("*"):
        if f.is_file():
            if f.name in ("model_metrics.json", "Stage3_Executive_Summary.txt",
                          "structural_fragility_assessment.txt", "targeting_rule.txt",
                          "model_equations.txt"):
                keep.append(("Stage3", f.name, "Final submission document"))

    return keep, discard


# ─────────────────────────────────────────────────────────────────────────────
#  STEP 5 — COMPLETENESS VALIDATION
# ─────────────────────────────────────────────────────────────────────────────

def validate_completeness(data):
    """Check all required submission items are present."""

    with open(STAGE3_OUT / "model_metrics.json", encoding="utf-8") as f:
        s3 = json.load(f)

    checks = []

    # 1. Total Expected Net Revenue
    val = s3.get("total_expected_net_revenue")
    checks.append(("Total Expected Net Revenue", "PRESENT" if val else "MISSING", f"{val:,.2f}" if val else ""))

    # 2. Targeted User Count
    val = s3.get("n_targeted_users")
    checks.append(("Targeted User Count", "PRESENT" if val else "MISSING", str(val) if val else ""))

    # 3. % Targeted
    val = s3.get("pct_targeted")
    checks.append(("% Targeted", "PRESENT" if val else "MISSING", f"{val}%" if val else ""))

    # 4. Return Exposure %
    val = s3.get("constraints", {}).get("return_risk", {}).get("actual")
    checks.append(("Return Exposure %", "PRESENT" if val else "MISSING", f"{val:.1%}" if val else ""))

    # 5. Top 10% Revenue Share
    val = s3.get("constraints", {}).get("concentration", {}).get("stage3_top10_share")
    checks.append(("Top 10% Revenue Share", "PRESENT" if val else "MISSING", f"{val:.2%}" if val else ""))

    # 6. Delta vs Baseline Concentration
    val = s3.get("constraints", {}).get("concentration", {}).get("relative_increase")
    checks.append(("Delta vs Baseline Concentration", "PRESENT" if val is not None else "MISSING", f"{val:.2%}" if val is not None else ""))

    # 7. Trade-off curves
    has_tradeoff = (STAGE3_CHR / "03_marginal_gain_budgets.png").exists()
    checks.append(("Trade-off Curves", "PRESENT" if has_tradeoff else "MISSING", "3 budget scenarios" if has_tradeoff else ""))

    # 8. Sensitivity results
    has_sensitivity = (STAGE3_TBL / "sensitivity_table.csv").exists()
    checks.append(("Sensitivity Results", "PRESENT" if has_sensitivity else "MISSING", "4 scenarios" if has_sensitivity else ""))

    # 9. Structural fragility
    has_fragility = (STAGE3_OUT / "structural_fragility_assessment.txt").exists()
    checks.append(("Structural Fragility Assessment", "PRESENT" if has_fragility else "MISSING", ""))

    # 10. Constraint compliance
    has_compliance = (STAGE3_TBL / "constraint_compliance.csv").exists()
    checks.append(("Constraint Compliance Dashboard", "PRESENT" if has_compliance else "MISSING", ""))

    # 11. Targeting rule
    has_rule = (STAGE3_OUT / "targeting_rule.txt").exists()
    checks.append(("Coherent Targeting Strategy", "PRESENT" if has_rule else "MISSING", ""))

    return checks


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 80)
    print("  FINAL SUBMISSION CONSOLIDATION — Case STABILIS")
    print("=" * 80)

    # === Build Excel workbook ===
    print("\n[1] BUILDING FINAL EXCEL WORKBOOK")
    data = build_workbook()

    # === Copy charts ===
    print("\n[2] SELECTING CHARTS FOR PPT")
    copied, missing = copy_charts()
    for c in copied:
        print(f"  [COPY] {c}")
    for m in missing:
        print(f"  [MISS] {m}")

    # === Copy key text documents ===
    print("\n[3] COPYING KEY DOCUMENTS")
    key_docs = [
        STAGE3_OUT / "Stage3_Executive_Summary.txt",
        STAGE3_OUT / "structural_fragility_assessment.txt",
        STAGE3_OUT / "targeting_rule.txt",
        STAGE3_OUT / "model_equations.txt",
        STAGE3_OUT / "model_metrics.json",
    ]
    for doc in key_docs:
        if doc.exists():
            shutil.copy2(str(doc), str(SUBMIT_DIR / doc.name))
            print(f"  [COPY] {doc.name}")

    # === File classification ===
    print("\n" + "=" * 80)
    print("  FILE CLASSIFICATION")
    print("=" * 80)
    keep, discard = classify_files()

    print("\n--- FILES TO KEEP (Final Submission) ---")
    for folder, name, reason in sorted(keep):
        print(f"  [KEEP]    {folder:30s} / {name:50s}  ({reason})")

    print(f"\n--- FILES TO DISCARD ({len(discard)} intermediate artifacts) ---")
    for folder, name, reason in sorted(discard):
        print(f"  [DISCARD] {folder:30s} / {name:50s}  ({reason})")

    # === Chart list for PPT ===
    print("\n" + "=" * 80)
    print("  FINAL CHART LIST FOR PPT")
    print("=" * 80)
    ppt_order = [
        ("Slide: Revenue Concentration",     "S1_lorenz_curve.png"),
        ("Slide: Revenue Concentration",     "S1_revenue_concentration.png"),
        ("Slide: Basket Distribution",        "S1_basket_log_distribution.png"),
        ("Slide: Model Calibration (S1)",     "S1_calibration_plot.png"),
        ("Slide: Validation Results",         "S2_calibration_curve.png"),
        ("Slide: Validation Results",         "S2_reliability_curve.png"),
        ("Slide: Model Validation",           "S2_gain_curve.png"),
        ("Slide: Optimization Trade-offs",    "S3_revenue_vs_return_risk.png"),
        ("Slide: Optimization Trade-offs",    "S3_revenue_vs_concentration.png"),
        ("Slide: Budget Sensitivity",         "S3_marginal_gain_budgets.png"),
        ("Slide: Targeting Distribution",     "S3_erpu_dist_targeted.png"),
        ("Slide: Constraint Compliance",      "S3_constraint_compliance.png"),
        ("Slide: Predictive Accuracy",        "S3_pred_vs_actual_erpu.png"),
    ]
    for i, (slide, chart) in enumerate(ppt_order, 1):
        status = "OK" if (SUBMIT_CHR / chart).exists() else "MISSING"
        print(f"  {i:2d}. [{status:4s}] {slide:40s}  {chart}")

    # === Completeness checklist ===
    print("\n" + "=" * 80)
    print("  STAGE 3 COMPLIANCE CHECKLIST")
    print("=" * 80)
    checks = validate_completeness(data)
    all_pass = True
    for item, status, value in checks:
        flag = "PASS" if status == "PRESENT" else "FAIL"
        if flag == "FAIL":
            all_pass = False
        val_str = f"  =  {value}" if value else ""
        print(f"  [{flag}] {item:40s}{val_str}")

    print("\n" + "-" * 80)
    if all_pass:
        print("  RESULT: ALL CHECKS PASSED — Submission package is COMPLETE")
    else:
        print("  RESULT: INCOMPLETE — Fix flagged items above")

    # === Final folder structure ===
    print("\n" + "=" * 80)
    print("  FINAL SUBMISSION FOLDER STRUCTURE")
    print("=" * 80)
    print(f"""
  final_submission/
  |-- Final_STABILIS_Submission.xlsx
  |     |-- Sheet: Stage1_Baseline
  |     |-- Sheet: Stage2_Validation
  |     |-- Sheet: Stage3_Test_Scoring
  |     |-- Sheet: Optimization_Summary
  |     |-- Sheet: Tradeoff_Table
  |     |-- Sheet: Sensitivity_Analysis
  |-- Stage3_Executive_Summary.txt
  |-- structural_fragility_assessment.txt
  |-- targeting_rule.txt
  |-- model_equations.txt
  |-- model_metrics.json
  |-- charts/
  |     |-- S1_lorenz_curve.png
  |     |-- S1_basket_log_distribution.png
  |     |-- S1_revenue_concentration.png
  |     |-- S1_calibration_plot.png
  |     |-- S2_calibration_curve.png
  |     |-- S2_reliability_curve.png
  |     |-- S2_gain_curve.png
  |     |-- S3_revenue_vs_return_risk.png
  |     |-- S3_revenue_vs_concentration.png
  |     |-- S3_marginal_gain_budgets.png
  |     |-- S3_erpu_dist_targeted.png
  |     |-- S3_constraint_compliance.png
  |     |-- S3_pred_vs_actual_erpu.png
""")

    print("=" * 80)
    print("  CONSOLIDATION COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    main()
